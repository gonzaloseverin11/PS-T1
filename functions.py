import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

# Inicializamos el DataFrame con los tipos de datos correctos
column_types = {
    'product_name': str,
    'description': str,
    'amount': int,
    'price': float,
    'category': str
}

df_products = pd.DataFrame({col: pd.Series(dtype=tipo) for col, tipo in column_types.items()})

def add_product(product_dict, df_products):
    """
    Añade un nuevo producto al DataFrame si no existe ya uno con el mismo nombre.

    Args:
    - product_dict (dict): Diccionario con los detalles del producto.
    - df_products (DataFrame): DataFrame donde se añadirá el producto.

    Returns:
    - DataFrame: El DataFrame actualizado.
    """
    product_name = product_dict.get('product_name', '').strip()

    if not product_name:
        print("El nombre del producto no puede estar vacío.")
        return df_products

    # Verificar si ya existe un producto con el mismo nombre (ignorando mayúsculas/minúsculas)
    if df_products['product_name'].str.lower().eq(product_name.lower()).any():
        print(f"El producto '{product_name}' ya existe. No se añadió.")
    else:
        df_product = pd.DataFrame([product_dict])
        df_products = pd.concat([df_products, df_product], ignore_index=True)
        print(f"Producto '{product_name}' añadido correctamente.")
    
    return df_products

# Captura de datos
product_dict = {
    'product_name': input("Introduce el nombre del producto: ").strip(),
    'description': input("Introduce la descripción del producto: ").strip(),
    'amount': int(input("Introduce la cantidad del producto: ")),
    'price': float(input("Introduce el precio del producto: ")),
    'category': input("Introduce la categoría del producto: ").strip()
}

df_products = add_product(product_dict, df_products)

def get_product(name_product, df_products):
    """
    Retrieves a product from the DataFrame based on its name.
    
    Parameters:
    name_product (str): The name of the product to retrieve.
    df_products (DataFrame): The DataFrame containing products.
    
    Returns:
    DataFrame: A DataFrame containing the product details if found, otherwise None.
    """
    # Filter the DataFrame for the product with the given name
    product = df_products[df_products['product_name'] == name_product]
    
    if not product.empty:
        return product
    else:
        return "El producto no existe en la base de datos."
    
name_product = input("Introduce el nombre del producto que deseas buscar: ")
get_product(name_product, df_products)

def delete_product(product, df_products):
    if product in df_products['product_name'].unique():
        df_products = df_products[df_products['product_name'] != product]
        print(f"El producto {product} ha sido eliminado de la base de datos.")
    else:
        print(f"El producto {product} no existe en la base de datos.")
    return df_products

product = input("Introduce el producto que deseas eliminar: ")
df_products = delete_product(product, df_products)

def edit_product(df_products):
    """
    Edita los detalles de un producto existente a partir de su nombre.

    Args:
    - df_products (DataFrame): DataFrame que contiene los productos.

    Returns:
    - DataFrame: El DataFrame actualizado.
    """
    product_name = input("Introduce el nombre del producto que deseas editar: ").strip().lower()

    # Buscar el índice del producto
    match = df_products['product_name'].str.lower() == product_name
    if not match.any():
        print(f"No se encontró ningún producto con el nombre '{product_name}'.")
        return df_products

    index = df_products[match].index[0]
    print("\nProducto actual:")
    print(df_products.loc[index])

    print("\nIntroduce los nuevos valores (deja en blanco para mantener el valor actual):")
    
    # Actualizar los campos si el usuario introduce un valor nuevo
    for column in ['description', 'amount', 'price', 'category']:
        current_value = df_products.at[index, column]
        new_value = input(f"{column.capitalize()} (actual: {current_value}): ").strip()
        
        if new_value:
            # Convertimos según el tipo de dato esperado
            if column == 'amount':
                df_products.at[index, column] = int(new_value)
            elif column == 'price':
                df_products.at[index, column] = float(new_value)
            else:
                df_products.at[index, column] = new_value

    print(f"\nProducto '{df_products.at[index, 'product_name']}' actualizado correctamente.")
    return df_products

df_products = edit_product(df_products)

producto = input("Que producto desea actualizar? ")
desicion = int(input("Deseas comprar o vender producto? 1. Comprar 2. Vender: "))
if desicion == 1:
    unidades = int(input("Cuantos productos desea comprar? "))
    df_products[df_products["product_name"] == producto]['amount'] += unidades
    print(f"Haz comprado {unidades} unidades de {producto}.")
elif desicion == 2:
    unidades = int(input("Cuantos productos desea vender? "))
    df_products[df_products["product_name"] == producto]['amount'] -= unidades
    print(f"Haz comprado {unidades} unidades de {producto}.")
else:
    print("Opcion no valida")

    product_name = input("Ingrese el nombre del producto a actualizar: ")
new_amount = int(input("Ingrese la nueva cantidad: "))

# Verificar si el producto existe
if product_name in df_products['product_name'].values:
    df_products.loc[df_products['product_name'] == product_name, 'amount'] = new_amount
    print("Cantidad actualizada con éxito!")
else:
    print("Producto no encontrado en el inventario.")


def generar_reporte_excel(df_products, excel_filename="reporte_productos.xlsx"):
    # Guardar el DataFrame en un archivo Excel
    df_products.to_excel(excel_filename, index=False, sheet_name="Productos")
    
    # Cargar el archivo para aplicar formato
    wb = load_workbook(excel_filename)
    ws = wb.active
    
    # Aplicar formato a los encabezados (negrita y centrado)
    for cell in ws[1]:  
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    # Ajustar ancho de columnas automáticamente
    column_widths = [max(len(str(value)) for value in df_products[col].astype(str)) for col in df_products.columns]
    
    for i, col_width in enumerate(column_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = col_width + 5  # Margen extra
    
    # Guardar el archivo con los cambios
    wb.save(excel_filename)
    
generar_reporte_excel(df_products)
